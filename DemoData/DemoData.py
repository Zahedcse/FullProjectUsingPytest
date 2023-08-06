import openpyxl


class HomePageData:
    data = [{"name": "Zahed", "email": "Zahed@mail.com", "password": "Bangladesh1#", "gender": "Male"},
            {"name": "Snigdha", "email": "Snigdha@mail.com", "password": "Bangladesh1#", "gender": "Female"},
            {"name": "Masud", "email": "Masud@mail.com", "password": "Bangladesh1#", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook('C:\\Users\\zahed\\Desktop\\new.xlsx')
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
